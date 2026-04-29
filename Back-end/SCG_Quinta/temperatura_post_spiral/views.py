from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from control_sala_cremas.models import ProductoSalaCremas
from .forms import RegistroTemperaturaPostSpiralForm
from .models import RegistroTemperaturaPostSpiral


def obtener_nombre_usuario(user):
    return getattr(user, 'nombre_completo', None) or user.get_full_name() or user.username


def aplicar_filtros_registros(request, queryset):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    turno = request.GET.get('turno')
    cliente = request.GET.get('cliente')
    producto = request.GET.get('producto')
    codigo = request.GET.get('codigo')
    lote = request.GET.get('lote')
    estado = request.GET.get('estado')

    if fecha_inicio:
        queryset = queryset.filter(fecha_hora__date__gte=fecha_inicio)

    if fecha_fin:
        queryset = queryset.filter(fecha_hora__date__lte=fecha_fin)

    if turno:
        queryset = queryset.filter(turno=turno)

    if cliente:
        queryset = queryset.filter(cliente__icontains=cliente)

    if producto:
        queryset = queryset.filter(producto__icontains=producto)

    if codigo:
        queryset = queryset.filter(codigo__icontains=codigo)

    if lote:
        queryset = queryset.filter(lote__icontains=lote)

    if estado == 'pendiente_accion':
        queryset = queryset.filter(
            acciones_correctivas_requieren_revision=True,
            acciones_correctivas_verificadas=False,
            verificado=False
        )

    elif estado == 'pendiente_final':
        queryset = queryset.filter(verificado=False).filter(
            Q(acciones_correctivas_requieren_revision=False) |
            Q(acciones_correctivas_verificadas=True)
        )

    elif estado == 'verificado':
        queryset = queryset.filter(verificado=True)

    return queryset


@login_required
def registrar_temperatura(request):
    if request.method == 'POST':
        form = RegistroTemperaturaPostSpiralForm(request.POST)

        if form.is_valid():
            registro = form.save(commit=False)
            registro.usuario = obtener_nombre_usuario(request.user)
            registro.save()

            messages.success(request, 'Registro guardado correctamente.')
            return redirect('temperatura_post_spiral:historial')
    else:
        form = RegistroTemperaturaPostSpiralForm()

    return render(request, 'temperatura_post_spiral/registrar.html', {
        'form': form
    })


@login_required
def historial(request):
    registros = RegistroTemperaturaPostSpiral.objects.all()
    registros = aplicar_filtros_registros(request, registros)

    clientes = (
        RegistroTemperaturaPostSpiral.objects
        .exclude(cliente__exact='')
        .values_list('cliente', flat=True)
        .distinct()
        .order_by('cliente')
    )

    productos = (
        RegistroTemperaturaPostSpiral.objects
        .exclude(producto__exact='')
        .values_list('producto', flat=True)
        .distinct()
        .order_by('producto')
    )

    return render(request, 'temperatura_post_spiral/historial.html', {
        'registros': registros,
        'clientes': clientes,
        'productos': productos,
        'filtros': request.GET
    })


@login_required
def detalle_registro(request, pk):
    registro = get_object_or_404(RegistroTemperaturaPostSpiral, pk=pk)

    return render(request, 'temperatura_post_spiral/detalle.html', {
        'registro': registro
    })


@login_required
def verificar_accion_correctiva(request, pk):
    if request.method != 'POST':
        return redirect('temperatura_post_spiral:historial')

    registro = get_object_or_404(RegistroTemperaturaPostSpiral, pk=pk)

    if registro.verificado:
        messages.warning(request, 'Este registro ya está verificado final.')
        return redirect('temperatura_post_spiral:detalle', pk=registro.pk)

    if not registro.acciones_correctivas_requieren_revision:
        messages.warning(request, 'Este registro no tiene acción correctiva para revisar.')
        return redirect('temperatura_post_spiral:detalle', pk=registro.pk)

    registro.acciones_correctivas_verificadas = True
    registro.fecha_revision_accion_correctiva = timezone.now()
    registro.nombre_revisor_accion_correctiva = obtener_nombre_usuario(request.user)
    registro.save(update_fields=[
        'acciones_correctivas_verificadas',
        'fecha_revision_accion_correctiva',
        'nombre_revisor_accion_correctiva',
        'actualizado_en'
    ])

    messages.success(request, 'Acción correctiva verificada correctamente.')
    return redirect('temperatura_post_spiral:detalle', pk=registro.pk)


@login_required
def verificar_final(request, pk):
    if request.method != 'POST':
        return redirect('temperatura_post_spiral:historial')

    registro = get_object_or_404(RegistroTemperaturaPostSpiral, pk=pk)

    if registro.verificado:
        messages.warning(request, 'Este registro ya está verificado final.')
        return redirect('temperatura_post_spiral:detalle', pk=registro.pk)

    if registro.acciones_correctivas_requieren_revision and not registro.acciones_correctivas_verificadas:
        messages.error(request, 'Primero debe verificarse la acción correctiva.')
        return redirect('temperatura_post_spiral:detalle', pk=registro.pk)

    registro.verificado = True
    registro.fecha_verificacion = timezone.now()
    registro.nombre_verificador = obtener_nombre_usuario(request.user)
    registro.save(update_fields=[
        'verificado',
        'fecha_verificacion',
        'nombre_verificador',
        'actualizado_en'
    ])

    messages.success(request, 'Registro verificado final correctamente.')
    return redirect('temperatura_post_spiral:detalle', pk=registro.pk)


@login_required
def api_productos_por_cliente(request):
    cliente = request.GET.get('cliente', '').strip()

    productos = ProductoSalaCremas.objects.filter(cliente=cliente).order_by('producto')

    data = []

    for p in productos:
        nombre_producto = getattr(p, 'producto', '') or getattr(p, 'nombre', '') or ''
        codigo = getattr(p, 'codigo', '') or ''

        data.append({
            'id': p.id,
            'producto': nombre_producto,
            'codigo': codigo,
        })

    return JsonResponse(data, safe=False)


@login_required
def descargar_excel(request):
    registros = RegistroTemperaturaPostSpiral.objects.all()
    registros = aplicar_filtros_registros(request, registros)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Temperatura Post Spiral'

    headers = [
        'Usuario',
        'Fecha y hora',
        'Turno',
        'Cliente',
        'Producto',
        'Código',
        'Lote',
        'Tiempo permanencia producto',
        'Temperatura °C',
        'Acción correctiva',
        'Observaciones',
        'Estado verificación',
        'Requiere revisión acción correctiva',
        'Acción correctiva verificada',
        'Fecha revisión acción correctiva',
        'Revisor acción correctiva',
        'Verificado final',
        'Fecha verificación final',
        'Verificador final',
        'Fecha creación',
        'Última actualización',
    ]

    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='B43C2C')
    header_font = Font(color='FFFFFF', bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r in registros:
        fecha_hora = timezone.localtime(r.fecha_hora).strftime('%d-%m-%Y %H:%M') if r.fecha_hora else ''
        fecha_revision = timezone.localtime(r.fecha_revision_accion_correctiva).strftime('%d-%m-%Y %H:%M') if r.fecha_revision_accion_correctiva else ''
        fecha_verificacion = timezone.localtime(r.fecha_verificacion).strftime('%d-%m-%Y %H:%M') if r.fecha_verificacion else ''
        creado_en = timezone.localtime(r.creado_en).strftime('%d-%m-%Y %H:%M') if r.creado_en else ''
        actualizado_en = timezone.localtime(r.actualizado_en).strftime('%d-%m-%Y %H:%M') if r.actualizado_en else ''

        ws.append([
            r.usuario,
            fecha_hora,
            r.turno,
            r.cliente,
            r.producto,
            r.codigo,
            r.lote,
            r.tiempo_permanencia_producto,
            float(r.temperatura) if r.temperatura is not None else '',
            r.accion_correctiva or '',
            r.observaciones or '',
            r.estado_verificacion,
            'Sí' if r.acciones_correctivas_requieren_revision else 'No',
            'Sí' if r.acciones_correctivas_verificadas else 'No',
            fecha_revision,
            r.nombre_revisor_accion_correctiva or '',
            'Sí' if r.verificado else 'No',
            fecha_verificacion,
            r.nombre_verificador or '',
            creado_en,
            actualizado_en,
        ])

    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = min(length + 3, 45)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="temperatura_post_spiral.xlsx"'

    wb.save(response)
    return response