from django.shortcuts import render
from django.http import JsonResponse, HttpResponseRedirect
from django.urls import reverse
from django.utils import timezone
from datetime import datetime, time
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
from django.db.models.functions import Trim, Upper
import json
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware

from .models import DatosFormularioControlDePesos, ProductoControlPeso

from functools import wraps
from urllib.parse import urlencode

from django.contrib import messages
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect
from django.views.decorators.http import require_POST

from .forms import ProductoControlPesoForm

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@login_required
def control_de_pesos(request):
    return render(request, 'control_de_pesos/r_control_de_pesos.html')


@csrf_exempt
@login_required
def vista_control_de_pesos(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'mensaje': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'mensaje': 'JSON inválido'}, status=400)

    dato = data.get('dato', None)
    if not dato:
        return JsonResponse({'ok': False, 'mensaje': 'No se recibió información'}, status=400)

    cliente = dato.get('cliente')
    codigo_producto = dato.get('codigo_producto')
    producto = dato.get('producto')
    peso_receta = dato.get('peso_receta')
    lote = dato.get('lote')
    turno = dato.get('turno')

    muestras = dato.get('muestras', [])

    # Compatibilidad por si viene el formato antiguo
    if not muestras and dato.get('peso_real') not in [None, ""]:
        muestras = [{
            'peso_real': dato.get('peso_real'),
            'altura': dato.get('altura')
        }]

    if not cliente or not producto or not peso_receta or not lote or not turno:
        return JsonResponse({
            'ok': False,
            'mensaje': 'Faltan datos obligatorios del encabezado'
        }, status=400)

    if not muestras:
        return JsonResponse({
            'ok': False,
            'mensaje': 'Debes ingresar al menos una muestra'
        }, status=400)

    nombre_tecnologo = getattr(request.user, 'nombre_completo', None) or request.user.username
    guardados = 0

    for muestra in muestras:
        peso_real = muestra.get('peso_real')
        altura = muestra.get('altura')

        if peso_real in [None, ""]:
            continue

        DatosFormularioControlDePesos.objects.create(
            nombre_tecnologo=nombre_tecnologo,
            fecha_registro=timezone.now(),
            cliente=cliente,
            codigo_producto=codigo_producto,
            producto=producto,
            peso_receta=int(peso_receta),
            peso_real=int(float(peso_real)),
            altura=float(altura) if altura not in [None, ""] else None,
            lote=lote,
            turno=turno
        )
        guardados += 1

    if guardados == 0:
        return JsonResponse({
            'ok': False,
            'mensaje': 'No se guardó ninguna muestra válida'
        }, status=400)

    return JsonResponse({
        'ok': True,
        'existe': True,
        'mensaje': f'Se guardaron {guardados} muestra(s) correctamente'
    })


@login_required
def redireccionar_selecciones_2(request):
    url_selecciones = reverse('vista_selecciones_2')
    return HttpResponseRedirect(url_selecciones)


@login_required
def graficos_control_pesos(request):
    clientes = DatosFormularioControlDePesos.objects.order_by().values_list('cliente', flat=True).distinct()
    turnos = DatosFormularioControlDePesos.objects.order_by().values_list('turno', flat=True).distinct()

    ctx = {
        'clientes': [c for c in clientes if c],
        'turnos': [t for t in turnos if t],
    }
    return render(request, 'control_de_pesos/graficos_control_pesos.html', ctx)


@login_required
@require_GET
def api_productos_por_cliente(request):
    cliente = (request.GET.get('cliente') or '').strip()
    if not cliente:
        return JsonResponse({'ok': True, 'productos': []})

    qs = (
        DatosFormularioControlDePesos.objects
        .annotate(cliente_norm=Upper(Trim('cliente')))
        .filter(cliente_norm__contains=cliente.upper())
        .order_by()
        .values('producto', 'codigo_producto')
        .distinct()
    )

    data = [
        {
            'producto': (r['producto'] or '').strip(),
            'codigo': r['codigo_producto']
        }
        for r in qs if (r['producto'] or '').strip()
    ]
    return JsonResponse({'ok': True, 'productos': data})


@login_required
@require_GET
def api_graficos_control_pesos(request):
    qs = (
        DatosFormularioControlDePesos.objects
        .annotate(
            cliente_norm=Upper(Trim('cliente')),
            producto_norm=Upper(Trim('producto')),
            turno_norm=Upper(Trim('turno')),
            lote_norm=Trim('lote'),
        )
    )

    cliente = (request.GET.get('cliente') or '').strip()
    producto = (request.GET.get('producto') or '').strip()
    turno = (request.GET.get('turno') or '').strip()
    lote = (request.GET.get('lote') or '').strip()
    desde = (request.GET.get('desde') or '').strip()
    hasta = (request.GET.get('hasta') or '').strip()

    if cliente:
        qs = qs.filter(cliente_norm__contains=cliente.upper())

    if producto:
        qs = qs.filter(producto_norm=producto.upper())

    if turno:
        qs = qs.filter(turno_norm=turno.upper())

    if lote:
        qs = qs.filter(lote_norm=lote)

    # -------------------------
    # FILTRO DESDE
    # -------------------------
    if desde:
        try:
            fecha_desde = parse_date(desde)

            if fecha_desde:
                dt_desde = make_aware(
                    datetime.combine(fecha_desde, time.min)
                )

                qs = qs.filter(fecha_registro__gte=dt_desde)

        except Exception as e:
            print("Error filtro desde:", e)

    # -------------------------
    # FILTRO HASTA
    # -------------------------
    if hasta:
        try:
            fecha_hasta = parse_date(hasta)

            if fecha_hasta:
                dt_hasta = make_aware(
                    datetime.combine(fecha_hasta, time.max)
                )

                qs = qs.filter(fecha_registro__lte=dt_hasta)

        except Exception as e:
            print("Error filtro hasta:", e)

    qs = qs.order_by('fecha_registro').values(
        'id',
        'nombre_tecnologo',
        'fecha_registro',
        'cliente',
        'producto',
        'codigo_producto',
        'peso_receta',
        'peso_real',
        'altura',
        'lote',
        'turno'
    )

    registros = []
    for r in qs:
        producto_base = ProductoControlPeso.objects.filter(
            area="TORTAS",
            cliente__iexact=r["cliente"],
            codigo=r["codigo_producto"],
            producto__iexact=r["producto"],
            activo=True
        ).first()

        peso_receta = int(producto_base.peso_receta) if producto_base else (
            int(r["peso_receta"]) if r["peso_receta"] is not None else None
        )

        perdida_operacional = float(producto_base.porcentaje_perdida) if producto_base else 0

        peso_maximo = None
        if peso_receta is not None and perdida_operacional < 100:
            peso_maximo = round(peso_receta / (1 - (perdida_operacional / 100)), 2)

        altura_objetivo = (
            float(producto_base.altura)
            if producto_base and producto_base.altura is not None
            else None
        )

        diferencia_altura = (
            float(producto_base.diff_altura)
            if producto_base and producto_base.diff_altura is not None
            else None
        )

        altura_minima = None
        altura_maxima = None

        if altura_objetivo is not None and diferencia_altura is not None:
            altura_minima = altura_objetivo - diferencia_altura
            altura_maxima = altura_objetivo + diferencia_altura
        peso_real = int(r["peso_real"]) if r["peso_real"] is not None else None
        altura_real = int(r["altura"]) if r["altura"] is not None else None

        desviacion = None

        if peso_real is not None and peso_receta is not None:
            if peso_maximo is not None:
                if peso_receta <= peso_real <= peso_maximo:
                    desviacion = 0
                elif peso_real < peso_receta:
                    desviacion = peso_real - peso_receta
                elif peso_real > peso_maximo:
                    desviacion = peso_real - peso_maximo
            else:
                desviacion = peso_real - peso_receta

        registros.append({
            "id": r["id"],
            "ts": r["fecha_registro"].isoformat(),
            "usuario": r["nombre_tecnologo"],
            "cliente": r["cliente"],
            "producto": r["producto"],
            "codigo_producto": r["codigo_producto"],

            "peso_receta": peso_receta,
            "peso_minimo": peso_receta,
            "peso_maximo": peso_maximo,
            "perdida_operacional": perdida_operacional,

            "peso_real": peso_real,
            "altura": altura_real,
            "altura_objetivo": altura_objetivo,
            "diferencia_altura": diferencia_altura,
            "altura_minima": altura_minima,
            "altura_maxima": altura_maxima,
            "desviacion": desviacion,

            "lote": r["lote"],
            "turno": r["turno"],
        })
    return JsonResponse({
        "ok": True,
        "registros": registros
    })


@login_required
def redireccionar_intermedio_4(request):
    url_intermedio = reverse('intermedio_4')
    return HttpResponseRedirect(url_intermedio)

@login_required
@require_GET
def api_clientes_control_pesos(request):
    area = (request.GET.get("area") or "TORTAS").strip().upper()

    clientes = (
        ProductoControlPeso.objects
        .filter(area=area, activo=True)
        .order_by("cliente")
        .values_list("cliente", flat=True)
        .distinct()
    )

    return JsonResponse({
        "ok": True,
        "clientes": list(clientes)
    })


@login_required
@require_GET
def api_productos_base_control_pesos(request):
    area = (request.GET.get("area") or "TORTAS").strip().upper()
    cliente = (request.GET.get("cliente") or "").strip()

    qs = ProductoControlPeso.objects.filter(area=area, activo=True)

    if cliente:
        qs = qs.filter(cliente__iexact=cliente)

    productos = list(
        qs.order_by("producto").values(
            "id",
            "cliente",
            "codigo",
            "producto",
            "peso_receta",
            "porcentaje_perdida",
            "altura",
            "diff_altura",
            "un_pp",
        )
    )

    return JsonResponse({
        "ok": True,
        "productos": productos
    })

def solo_usuarios_autorizados(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        usuarios_autorizados = {1, 49, 58}

        if request.user.pk not in usuarios_autorizados:
            raise PermissionDenied(
                "No tienes autorización para administrar los productos."
            )

        return view_func(request, *args, **kwargs)

    return wrapper


@solo_usuarios_autorizados
def administrar_productos_control_peso(request):
    producto_id = request.GET.get("editar") or request.POST.get("producto_id")
    producto_edicion = None

    if producto_id:
        producto_edicion = get_object_or_404(
            ProductoControlPeso,
            pk=producto_id,
        )

    if request.method == "POST":
        formulario = ProductoControlPesoForm(
            request.POST,
            instance=producto_edicion,
        )

        if formulario.is_valid():
            producto_guardado = formulario.save()

            if producto_edicion:
                messages.success(
                    request,
                    f'El producto "{producto_guardado.producto}" '
                    "fue actualizado correctamente."
                )
            else:
                messages.success(
                    request,
                    f'El producto "{producto_guardado.producto}" '
                    "fue creado correctamente."
                )

            parametros_retorno = {
                "buscar": request.POST.get("buscar", "").strip(),
                "area": request.POST.get("area", "").strip(),
                "cliente": request.POST.get("cliente", "").strip(),
                "estado": request.POST.get("estado", "").strip(),
                "page": request.POST.get("page", "").strip(),
            }

            parametros_retorno = {
                clave: valor
                for clave, valor in parametros_retorno.items()
                if valor
            }

            url_retorno = reverse("administrar_productos_control_peso")

            if parametros_retorno:
                url_retorno += f"?{urlencode(parametros_retorno)}"

            return redirect(url_retorno)

        messages.error(
            request,
            "No fue posible guardar el producto. Revisa los campos indicados."
        )

    else:
        formulario = ProductoControlPesoForm(
            instance=producto_edicion
        )

    productos = ProductoControlPeso.objects.all()

    busqueda = request.GET.get("buscar", "").strip()
    area = request.GET.get("area", "").strip()
    cliente = request.GET.get("cliente", "").strip()
    estado = request.GET.get("estado", "").strip()

    if busqueda:
        productos = productos.filter(
            Q(codigo__icontains=busqueda)
            | Q(producto__icontains=busqueda)
            | Q(cliente__icontains=busqueda)
        )

    if area:
        productos = productos.filter(area=area)

    if cliente:
        productos = productos.filter(cliente=cliente)

    if estado == "activos":
        productos = productos.filter(activo=True)
    elif estado == "inactivos":
        productos = productos.filter(activo=False)

    productos = productos.order_by(
        "area",
        "cliente",
        "producto",
    )

    paginador = Paginator(productos, 20)
    pagina = paginador.get_page(request.GET.get("page"))

    filtros_paginacion = {
        "buscar": busqueda,
        "area": area,
        "cliente": cliente,
        "estado": estado,
    }

    filtros_paginacion = {
        clave: valor
        for clave, valor in filtros_paginacion.items()
        if valor
    }

    contexto = {
        "formulario": formulario,
        "producto_edicion": producto_edicion,
        "pagina": pagina,
        "busqueda": busqueda,
        "area_seleccionada": area,
        "cliente_seleccionado": cliente,
        "estado_seleccionado": estado,
        "areas": ProductoControlPeso.AREA_CHOICES,
        "clientes": ProductoControlPeso.CLIENTE_CHOICES,
        "filtros_qs": urlencode(filtros_paginacion),
    }

    return render(
        request,
        "control_de_pesos/administrar_productos.html",
        contexto,
    )


@solo_usuarios_autorizados
@require_POST
def cambiar_estado_producto_control_peso(request, producto_id):
    producto = get_object_or_404(
        ProductoControlPeso,
        pk=producto_id,
    )

    producto.activo = not producto.activo
    producto.save(update_fields=["activo"])

    estado = "activado" if producto.activo else "desactivado"

    messages.success(
        request,
        f'El producto "{producto.producto}" fue {estado}.'
    )

    parametros_retorno = {
        "buscar": request.POST.get("buscar", "").strip(),
        "area": request.POST.get("area", "").strip(),
        "cliente": request.POST.get("cliente", "").strip(),
        "estado": request.POST.get("estado", "").strip(),
        "page": request.POST.get("page", "").strip(),
    }

    parametros_retorno = {
        clave: valor
        for clave, valor in parametros_retorno.items()
        if valor
    }

    url_retorno = reverse("administrar_productos_control_peso")

    if parametros_retorno:
        url_retorno += f"?{urlencode(parametros_retorno)}"

    return redirect(url_retorno)

@login_required
def redireccionar_index(request):
    url_index = reverse('index')
    return HttpResponseRedirect(url_index)

@solo_usuarios_autorizados
def descargar_productos_control_peso_excel(request):
    productos = ProductoControlPeso.objects.all()

    busqueda = request.GET.get("buscar", "").strip()
    area = request.GET.get("area", "").strip()
    cliente = request.GET.get("cliente", "").strip()
    estado = request.GET.get("estado", "").strip()

    # Aplicar los mismos filtros del template
    if busqueda:
        productos = productos.filter(
            Q(codigo__icontains=busqueda)
            | Q(producto__icontains=busqueda)
            | Q(cliente__icontains=busqueda)
        )

    if area:
        productos = productos.filter(area=area)

    if cliente:
        productos = productos.filter(cliente=cliente)

    if estado == "activos":
        productos = productos.filter(activo=True)

    elif estado == "inactivos":
        productos = productos.filter(activo=False)

    productos = productos.order_by(
        "area",
        "cliente",
        "producto",
    )

    # Crear archivo Excel
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Productos"

    encabezados = [
        "ID",
        "Área",
        "Cliente",
        "Código",
        "Producto",
        "Peso receta (gr)",
        "Pérdida operacional (%)",
        "Peso máximo calculado (gr)",
        "Altura objetivo (mm)",
        "Diferencia de altura (mm)",
        "Altura mínima (mm)",
        "Altura máxima (mm)",
        "Unidades por persona",
        "Estado",
    ]

    hoja.append(encabezados)

    # Formato del encabezado
    relleno_encabezado = PatternFill(
        fill_type="solid",
        fgColor="B43C2C",
    )

    for celda in hoja[1]:
        celda.fill = relleno_encabezado
        celda.font = Font(
            color="FFFFFF",
            bold=True,
        )
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for producto in productos:
        porcentaje_perdida = float(
            producto.porcentaje_perdida or 0
        )

        peso_maximo = None

        if porcentaje_perdida < 100:
            peso_maximo = round(
                producto.peso_receta
                / (1 - porcentaje_perdida / 100),
                2,
            )

        altura_objetivo = producto.altura

        diferencia_altura = float(
            producto.diff_altura or 0
        )

        altura_minima = None
        altura_maxima = None

        if altura_objetivo is not None:
            altura_minima = round(
                float(altura_objetivo) - diferencia_altura,
                2,
            )

            altura_maxima = round(
                float(altura_objetivo) + diferencia_altura,
                2,
            )

        hoja.append([
            producto.id,
            producto.get_area_display(),
            producto.get_cliente_display(),
            producto.codigo,
            producto.producto,
            producto.peso_receta,
            porcentaje_perdida,
            peso_maximo,
            altura_objetivo,
            diferencia_altura,
            altura_minima,
            altura_maxima,
            (
                float(producto.un_pp)
                if producto.un_pp is not None
                else None
            ),
            "Activo" if producto.activo else "Inactivo",
        ])

    # Inmovilizar encabezado
    hoja.freeze_panes = "A2"

    # Activar filtros en Excel
    hoja.auto_filter.ref = hoja.dimensions

    # Alinear contenido
    for fila in hoja.iter_rows(
        min_row=2,
        max_row=hoja.max_row,
    ):
        for celda in fila:
            celda.alignment = Alignment(
                vertical="center"
            )

    # Anchos de columnas
    anchos = {
        1: 10,   # ID
        2: 20,   # Área
        3: 22,   # Cliente
        4: 16,   # Código
        5: 42,   # Producto
        6: 20,   # Peso receta
        7: 24,   # Pérdida
        8: 28,   # Peso máximo
        9: 23,   # Altura objetivo
        10: 27,  # Diferencia altura
        11: 22,  # Altura mínima
        12: 22,  # Altura máxima
        13: 24,  # Unidades persona
        14: 15,  # Estado
    }

    for numero_columna, ancho in anchos.items():
        letra = get_column_letter(numero_columna)
        hoja.column_dimensions[letra].width = ancho

    # Formato numérico con dos decimales
    columnas_decimales = ["G", "H", "J", "K", "L", "M"]

    for columna in columnas_decimales:
        for celda in hoja[columna][1:]:
            celda.number_format = "0.00"

    fecha = timezone.localdate().strftime("%Y-%m-%d")
    nombre_archivo = f"productos_control_peso_{fecha}.xlsx"

    respuesta = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    respuesta["Content-Disposition"] = (
        f'attachment; filename="{nombre_archivo}"'
    )

    libro.save(respuesta)

    return respuesta